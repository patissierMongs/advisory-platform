"""엔드투엔드 스모크 테스트 — 전체 파이프라인 + 게이트 검증.

실행: .venv/Scripts/python.exe smoke_test.py
임시 데이터 디렉터리를 사용하므로 실제 data/ 를 건드리지 않는다.
"""
import os
import tempfile
import io

TMP = tempfile.mkdtemp(prefix="advisory_smoke_")
os.environ["ADVISORY_DATA_DIR"] = TMP
os.environ["ADVISORY_DATABASE_URL"] = f"sqlite:///{os.path.join(TMP, 'test.db')}"
os.environ["ADVISORY_BUNDLED_FEEDS"] = "false"  # 결정적 테스트 — 동봉 CVE 대량적재 비활성

from fastapi.testclient import TestClient  # noqa: E402
from app.main import app  # noqa: E402
from app.seed import _minimal_pdf  # noqa: E402

ok = 0
fail = 0


def check(name, cond, extra=""):
    global ok, fail
    if cond:
        ok += 1
        print(f"  PASS  {name}")
    else:
        fail += 1
        print(f"  FAIL  {name}  {extra}")


with TestClient(app) as c:
    # health
    check("health", c.get("/api/health").json()["status"] == "ok")

    # dashboard
    dash = c.get("/api/v1/dashboard").json()
    check("dashboard stats", len(dash["stats"]) >= 3, dash["stats"])
    check("dashboard not_found=2", dash["stats"][1]["value"] == 2, dash["stats"][1])

    # active advisory
    advs = c.get("/api/v1/advisories").json()["items"]
    active = next(a for a in advs if a["status"] == "NEEDS_CVE_UPDATE")
    aid = active["id"]
    check("advisory seeded", active["doc_no"] == "국정원-사이버-2026-0612")

    cves = c.get(f"/api/v1/advisories/{aid}/cves").json()
    check("6 extracted", cves["summary"]["extracted"] == 6, cves["summary"])
    check("4 found / 2 not_found", cves["summary"]["found"] == 4 and cves["summary"]["not_found"] == 2)
    check("gate closed", cves["can_proceed"] is False)

    # match must be blocked by gate (409)
    r = c.post(f"/api/v1/advisories/{aid}/match")
    check("match blocked (409)", r.status_code == 409, r.status_code)

    # upload + apply feed
    feed_path = os.path.join(os.path.dirname(__file__), "samples", "krcert_cve_feed_2026-06-15.json")
    with open(feed_path, "rb") as f:
        r = c.post("/api/v1/cve-feeds", files={"file": ("krcert_cve_feed_2026-06-15.json", f, "application/json")})
    fimp = r.json()
    check("feed validated +2", fimp["status"] == "VALIDATED" and fimp["added_count"] == 2, fimp)
    r = c.post(f"/api/v1/cve-feeds/{fimp['import_id']}/apply").json()
    check("feed applied, 1 unlocked", r["added_count"] == 2 and r["advisories_unlocked"] == 1, r)

    # gate now open
    cves2 = c.get(f"/api/v1/advisories/{aid}/cves").json()
    check("gate open", cves2["can_proceed"] is True)
    check("cve db count grew", c.get("/api/v1/cves/stats").json()["count"] == 6)

    # match
    m = c.post(f"/api/v1/advisories/{aid}/match").json()
    check("match produced rows", m["matched"] > 0, m)
    matches = c.get(f"/api/v1/advisories/{aid}/matches").json()
    check("matches listed", matches["summary"]["active"] == m["matched"], matches["summary"])
    # version comparator sanity: Chrome 122.x/121.x < 124 → matched
    chrome = [x for x in matches["items"] if x["cve"] == "CVE-2026-30012"]
    check("chrome lt-rule matched 2", len(chrome) == 2, chrome)
    # Windows 10 must NOT match windows_11 CVE
    win = [x for x in matches["items"] if x["cve"] == "CVE-2026-21345"]
    check("win11 only (no win10)", all("PC-0301" != x["asset_no"] for x in win), win)

    # exclude a match (false positive)
    first = matches["items"][0]
    r = c.patch(f"/api/v1/matches/{first['id']}", json={"status": "EXCLUDED", "reason": "오탐 검토"})
    check("exclude ok", r.json()["status"] == "EXCLUDED")

    # notification preview + send
    prev = c.get(f"/api/v1/advisories/{aid}/notification-preview").json()
    check("preview has departments", prev["total_departments"] > 0, prev["total_departments"])
    check("message generated", "[보안조치 요청]" in prev["departments"][0]["message"])
    send = c.post(f"/api/v1/advisories/{aid}/notifications", json={"all": True}).json()
    check("send all SENT", all(r2["status"] == "SENT" for r2 in send["results"]), send)
    # idempotency: re-send returns idempotent
    send2 = c.post(f"/api/v1/advisories/{aid}/notifications", json={"all": True}).json()
    check("idempotent re-send", all(r2.get("idempotent") for r2 in send2["results"]), send2)

    hist = c.get("/api/v1/notifications").json()["items"]
    check("history populated", len(hist) >= 5, len(hist))

    # advisory should be COMPLETED now
    check("advisory completed", c.get(f"/api/v1/advisories/{aid}").json()["status"] == "COMPLETED")

    # PDF upload + extract
    pdf = _minimal_pdf(["Test advisory", "CVE-2026-21345 and CVE-2026-21401"])
    r = c.post("/api/v1/advisories", files={"file": ("test.pdf", io.BytesIO(pdf), "application/pdf")},
               data={"source_org": "국가정보원", "receive_channel": "NCST", "doc_no": "TEST-001"})
    check("pdf upload 201", r.status_code == 201, r.status_code)
    new_id = r.json()["id"]
    # duplicate detection
    r = c.post("/api/v1/advisories", files={"file": ("test.pdf", io.BytesIO(pdf), "application/pdf")},
               data={"source_org": "국가정보원", "receive_channel": "NCST"})
    check("duplicate 409", r.status_code == 409, r.status_code)
    disp = c.post(f"/api/v1/advisories/{new_id}/extract").json()
    check("extract dispatched(async, queued)", disp.get("extract_phase") == "queued", disp)
    import time as _t
    for _ in range(60):  # 비동기 추출 완료 대기(백그라운드 워커)
        if c.get(f"/api/v1/advisories/{new_id}").json().get("extract_phase") in ("done", "failed"):
            break
        _t.sleep(0.1)
    ext = c.get(f"/api/v1/advisories/{new_id}/cves").json()["summary"]
    check("extract found 2 cves(async done)", ext["extracted"] == 2 and ext["found"] == 2, ext)

    # asset import (generate xlsx)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["자산번호", "사용부서", "운영체제/SW", "세부버전", "IP", "담당자"])
    ws.append(["PC-9001", "도로국", "Windows 11", "23H2", "10.20.5.99", "테스터"])
    ws.append(["PC-9002", "없는부서", "Windows 11", "22H2", "", "미상"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = c.post("/api/v1/assets/import/preview",
               files={"file": ("assets.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    pv = r.json()
    check("preview columns", len(pv["columns"]) == 6, pv.get("columns"))
    check("suggested mapping", pv["suggested_mapping"].get("department") == "B", pv["suggested_mapping"])
    r = c.post(f"/api/v1/assets/import/{pv['import_id']}/commit",
               json={"mapping": pv["suggested_mapping"], "mode": "append", "on_warning": "skip"}).json()
    check("commit 2 (없는부서 자동생성)", r["committed"] == 2 and "없는부서" in r.get("created_departments", []), r)

    # ========== 발송 후 단계 (§★★★★★ ~ ★★★) ==========
    notifs = c.get("/api/v1/notifications").json()["items"]
    target = next(n for n in notifs if n["advisory_id"] == aid)
    nid = target["id"]
    # ── 조치 회신 루프 ──
    r = c.patch(f"/api/v1/notifications/{nid}/ack", json={"ack_status": "IN_PROGRESS", "note": "패치 적용 중", "by": "김담당"})
    check("ack 진행중+담당", r.json()["ack_status"] == "IN_PROGRESS" and r.json()["ack_by"] == "김담당", r.json())
    r = c.patch(f"/api/v1/notifications/{nid}/ack", json={"ack_status": "UNABLE"})
    check("조치불가 사유필수(400)", r.status_code == 400, r.status_code)
    r = c.patch(f"/api/v1/notifications/{nid}/ack", json={"ack_status": "DONE", "note": "완료"})
    check("ack 완료", r.json()["ack_status"] == "DONE")
    r = c.post(f"/api/v1/notifications/{nid}/evidence",
               files={"file": ("patch.png", io.BytesIO(b"\x89PNG evidence"), "image/png")})
    check("증빙 업로드", r.json()["evidence"] == "patch.png", r.json())
    # ── 보안: 파일명 sanitize(경로 traversal 차단) ──
    from app.core.files import safe_filename
    check("safe_filename 기본명만", safe_filename("a/../../evil.txt") == "evil.txt"
          and safe_filename("..") == "file" and safe_filename("보고서.pdf") == "보고서.pdf")
    ev_dir = os.path.join(TMP, "evidence")
    r = c.post(f"/api/v1/notifications/{nid}/evidence",
               files={"file": ("x/../../_SHOULD_NOT_ESCAPE.txt", io.BytesIO(b"x"), "text/plain")})
    escaped = any(os.path.exists(os.path.join(d, "_SHOULD_NOT_ESCAPE.txt"))
                  for d in (TMP, os.path.dirname(TMP), os.path.dirname(os.path.dirname(TMP))))
    inside = any("_SHOULD_NOT_ESCAPE.txt" in f for f in os.listdir(ev_dir))
    check("traversal 차단(EVIDENCE_DIR 안에만 기록)", r.status_code == 200 and not escaped and inside,
          f"http={r.status_code} escaped={escaped} inside={inside}")
    # ── 보안: 증빙 업로드 크기 제한(413) ──
    from app.config import settings as _st
    _old_mb = _st.MAX_UPLOAD_MB
    _st.MAX_UPLOAD_MB = 0
    r = c.post(f"/api/v1/notifications/{nid}/evidence",
               files={"file": ("big.bin", io.BytesIO(b"AB"), "application/octet-stream")})
    check("증빙 크기 제한(413)", r.status_code == 413, r.status_code)
    _st.MAX_UPLOAD_MB = _old_mb
    rem = c.get(f"/api/v1/advisories/{aid}/remediation").json()
    check("조치율 산출", rem["dept_total"] > 0 and rem["done"] >= 1, rem.get("done_rate"))
    # ── 보고서 자동 생성 ──
    r = c.get(f"/api/v1/advisories/{aid}/report.xlsx")
    check("Excel 보고서", r.status_code == 200 and r.content[:2] == b"PK", r.status_code)
    r = c.get(f"/api/v1/advisories/{aid}/report.html")
    check("HTML 보고서(한글)", r.status_code == 200 and "조치 결과 보고서" in r.text)
    # ── SLA / 리마인드 ──
    r = c.post(f"/api/v1/advisories/{aid}/remind", json={}).json()
    check("미회신 리마인드", r["reminded"] >= 1, r)
    check("리마인드 대상 조회", "items" in c.get("/api/v1/reminders/due").json())
    # ── 추출 CVE 수동 추가/삭제(게이트 재평가) ──
    r = c.post(f"/api/v1/advisories/{aid}/cves", json={"cve_id": "CVE 2026 99999"})
    check("CVE 수동추가 201(공백교정)", r.status_code == 201 and r.json()["cve_id_text"] == "CVE-2026-99999", r.status_code)
    new_ac = r.json()["id"]
    check("미등록 추가→게이트 닫힘", c.get(f"/api/v1/advisories/{aid}/cves").json()["can_proceed"] is False)
    c.delete(f"/api/v1/advisory-cves/{new_ac}")
    check("삭제→게이트 열림", c.get(f"/api/v1/advisories/{aid}/cves").json()["can_proceed"] is True)
    # ── 오탐 제외 기억 ──
    rules = c.get("/api/v1/exclusion-rules").json()["items"]
    check("오탐 규칙 기억", len(rules) >= 1, len(rules))
    rerun = c.post(f"/api/v1/advisories/{aid}/match").json()
    check("재매칭 제외제안 count", rerun.get("suggested_exclude", 0) >= 1, rerun)
    ms = c.get(f"/api/v1/advisories/{aid}/matches").json()["items"]
    check("매칭에 suggested_exclude", any(m["suggested_exclude"] for m in ms))
    # ── 그룹웨어 게시판 + 웹훅 ack ──
    r = c.post(f"/api/v1/advisories/{aid}/board").json()
    check("게시판 게시", str(r.get("board_post_id", "")).startswith("BOARD-"), r)
    r = c.post("/api/v1/webhooks/groupware/ack", json={"department": "도로국", "status": "완료", "by": "댓글회신"})
    check("게시판 회신→ack 동기화", r.status_code == 200 and r.json()["ack_status"] == "DONE", r.status_code)
    # ── 대시보드 SLA ──
    dash2 = c.get("/api/v1/dashboard").json()
    check("대시보드 SLA+리마인드", "sla" in dash2 and "due_reminders" in dash2, list(dash2.keys()))

    # ========== 엑셀: 제목행 + 세로 병합 부서 (현실 자산대장) ==========
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["○○부 자산관리대장", None, None, None, None, None])              # 1행: 제목(병합)
    ws2.append(["자산번호", "사용부서", "운영체제/SW", "세부버전", "IP", "담당자"])  # 2행: 실제 헤더
    ws2.append(["PC-2001", "정보화담당관실", "Windows 11", "23H2", "10.30.1.1", "박하나"])
    ws2.append(["PC-2002", None, "Windows 11", "22H2", "10.30.1.2", "박두리"])      # 부서 세로병합
    ws2.append(["PC-2003", "도로국", "Windows Server", "2022", "10.30.2.1", "이세찬"])
    ws2.append(["PC-2004", None, "Microsoft Office", "2021", "", "이네찬"])         # 부서 세로병합
    ws2.append(["PC-2005", None, "Google Chrome", "122.x", "", "이다찬"])           # 부서 세로병합
    ws2.merge_cells("A1:F1")
    ws2.merge_cells("B3:B4")   # 정보화담당관실 (행3~4)
    ws2.merge_cells("B5:B7")   # 도로국 (행5~7)
    buf2 = io.BytesIO()
    wb2.save(buf2)
    buf2.seek(0)
    r = c.post("/api/v1/assets/import/preview",
               files={"file": ("messy.xlsx", buf2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    pv2 = r.json()
    check("제목행 무시·헤더행 자동감지=2행", pv2["detected_header_row"] == 2, pv2.get("detected_header_row"))
    check("헤더 인식→부서=B 매핑", pv2["suggested_mapping"].get("department") == "B", pv2["suggested_mapping"])
    check("헤더 선택용 후보행 반환", len(pv2.get("preview_rows", [])) >= 7)
    r = c.post(f"/api/v1/assets/import/{pv2['import_id']}/commit",
               json={"mapping": pv2["suggested_mapping"], "header_row": 2, "mode": "append", "on_warning": "skip"}).json()
    dept_warn = [w for w in r["warnings"] if "DEPARTMENT" in w["issue"]]
    check("세로병합 부서 채움→5행 전부 적재(누락 0)", r["committed"] == 5 and len(dept_warn) == 0, r)
    a = c.get("/api/v1/assets?q=PC-2005").json()["items"]
    check("병합 마지막행 부서=도로국(채워짐)", bool(a) and a[0]["department"] == "도로국", a)

    # ========== 엑셀: 2층(다단) 헤더 결합 ==========
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.append(["자산번호", "사용부서", "제품", None, "담당자"])   # 1층: 그룹 라벨
    ws3.append([None, None, "OS", "버전", None])                  # 2층: 세부 라벨
    ws3.append(["PC-3001", "정보화담당관실", "Windows 11", "23H2", "담당가"])
    ws3.append(["PC-3002", "도로국", "Windows Server", "2022", "담당나"])
    ws3.merge_cells("C1:D1")  # "제품" 이 OS·버전 두 칸 위에 걸침
    ws3.merge_cells("A1:A2"); ws3.merge_cells("B1:B2"); ws3.merge_cells("E1:E2")  # 세로 병합 헤더
    buf3 = io.BytesIO()
    wb3.save(buf3)
    buf3.seek(0)
    r = c.post("/api/v1/assets/import/preview",
               data={"header_row": 1, "header_rows": 2},
               files={"file": ("multi.xlsx", buf3, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    pv3 = r.json()
    hdrs = [col["header"] for col in pv3["columns"]]
    check("다단 헤더 결합('제품 / OS')", "제품 / OS" in hdrs and "제품 / 버전" in hdrs, hdrs)
    check("결합 헤더→제품=C·버전=D 매핑",
          pv3["suggested_mapping"].get("product_key") == "C" and pv3["suggested_mapping"].get("version_norm") == "D",
          pv3["suggested_mapping"])
    r = c.post(f"/api/v1/assets/import/{pv3['import_id']}/commit",
               json={"mapping": pv3["suggested_mapping"], "header_row": 1, "header_rows": 2,
                     "mode": "append", "on_warning": "skip"}).json()
    check("2층 헤더 후 2행 적재", r["committed"] == 2, r)

    # ========== 엑셀: 한 셀 다중값 → 구분자 분할 매핑 ==========
    wb4 = Workbook()
    ws4 = wb4.active
    ws4.append(["자산번호", "사용부서", "운영체제", "담당"])
    ws4.append(["PC-7001", "총무국", "Windows 11", "좋은팀, 홍길동, 010-1234-5678"])
    buf4 = io.BytesIO(); wb4.save(buf4); buf4.seek(0)
    pv4 = c.post("/api/v1/assets/import/preview",
                 files={"file": ("split.xlsx", buf4, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}).json()
    split_map = {"asset_no": "A", "department": "B", "product_key": "C",
                 "owner_team": {"col": "D", "sep": ",", "part": 0},
                 "owner_name": {"col": "D", "sep": ",", "part": 1},
                 "owner_contact": {"col": "D", "sep": ",", "part": 2}}
    r = c.post(f"/api/v1/assets/import/{pv4['import_id']}/commit",
               json={"mapping": split_map, "mode": "append", "on_warning": "skip"}).json()
    check("분할 매핑 적재", r["committed"] == 1, r)
    a7 = c.get("/api/v1/assets", params={"q": "PC-7001"}).json()["items"]
    a7 = a7[0] if a7 else {}
    check("담당 셀 분할: 팀/이름/연락처",
          a7.get("owner_team") == "좋은팀" and a7.get("owner_name") == "홍길동"
          and a7.get("owner_contact") == "010-1234-5678", a7)

    # ========== STEP2 원문 뷰어: 페이지 렌더 + CVE 강조 박스 ==========
    view = c.get(f"/api/v1/advisories/{aid}/pdf-view").json()
    check("pdf-view available + 강조박스", view.get("available") is True and len(view.get("boxes", [])) >= 1, view)
    rp = c.get(f"/api/v1/advisories/{aid}/page/0.png")
    check("페이지 PNG 200·image/png", rp.status_code == 200 and rp.headers["content-type"] == "image/png", rp.status_code)
    rdl = c.get(f"/api/v1/advisories/{aid}/file", params={"download": 1})
    check("원문 다운로드 attachment", "attachment" in rdl.headers.get("content-disposition", ""), rdl.headers.get("content-disposition"))

    # ========== 활동 기록: 감사 로그 조회 ==========
    au = c.get("/api/v1/audit").json()
    check("활동 기록 조회", au["total"] >= 1 and bool(au["items"]) and "action_ko" in au["items"][0], au.get("total"))
    upl = c.get("/api/v1/audit", params={"action": "ADVISORY_UPLOAD"}).json()["items"]
    check("업로드 기록·원문 연결(has_pdf)", bool(upl) and any(i.get("has_pdf") for i in upl), len(upl))

print(f"\n=== {ok} passed, {fail} failed ===")
raise SystemExit(1 if fail else 0)
