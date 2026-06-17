"""자산 엑셀 가져오기 — 미리보기·추천매핑·커밋 (명세서 §4.5)."""
from __future__ import annotations

from sqlalchemy import select
from sqlalchemy.orm import Session

from .. import enums
from ..models import Asset, Department
from .normalize import normalize_product, split_product_version
from .versioning import normalize_version

# 시스템 필드 → 헤더 별칭(부분일치, 소문자)
FIELD_HINTS: dict[str, list[str]] = {
    "asset_no": ["자산번호", "자산no", "자산 번호", "asset", "관리번호", "번호"],
    "department": ["부서", "사용부서", "담당부서", "dept", "department"],
    "product_key": ["운영체제", "os", "제품", "sw", "소프트웨어", "product", "프로그램"],
    "version_norm": ["버전", "version", "세부버전", "ver"],
    "ip": ["ip", "아이피", "주소"],
    "owner_name": ["담당자", "사용자", "owner", "담당", "이름"],
    "owner_team": ["담당팀", "소속팀", "소속"],
    "owner_contact": ["연락처", "전화", "휴대폰", "핸드폰", "전화번호", "contact", "phone", "tel"],
}
REQUIRED_FIELDS = ("department", "product_key")  # 버전은 선택 — 제품 셀에서 분리하거나 '*' 규칙으로 매칭


def col_letter(idx0: int) -> str:
    """0-기반 컬럼 인덱스 → 엑셀 레터(A, B, … Z, AA)."""
    s = ""
    n = idx0 + 1
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def letter_to_idx(letter: str) -> int:
    n = 0
    for ch in letter.strip().upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def parse_spec(value) -> tuple[int, str, int | None, str] | None:
    """매핑 값 정규화 → (컬럼인덱스, 구분자, 파트, 모드). 한 셀에 여러 값이 든 경우 지원.

    value 는 다음 중 하나:
      · 컬럼 레터(str, 예 "F")                         → 셀 전체
      · {"col":"F","sep":",","part":1}                 → 구분자로 나눈 part 번째 조각(ScanOps 패턴)
      · {"col":"F","mode":"product"|"version"}         → 제품·버전 자동 분리(예 "Windows 11 22H2")
    """
    if isinstance(value, dict):
        letter = (value.get("col") or value.get("letter") or "").strip()
        sep = value.get("sep") or ""
        part = value.get("part")
        part = part if isinstance(part, int) else None
        mode = value.get("mode") or ""
    else:
        letter, sep, part, mode = str(value or "").strip(), "", None, ""
    if not letter:
        return None
    return letter_to_idx(letter), sep, part, mode


def resolve_cell(row: list, spec: tuple[int, str, int | None, str] | None) -> str | None:
    """행에서 spec 에 따라 셀 값 추출. 구분자 분할 / 제품·버전 자동분리 지원."""
    if not spec:
        return None
    i, sep, part, mode = spec
    if i is None or i < 0 or i >= len(row):
        return None
    v = row[i]
    if v is None:
        return None
    text = str(v).strip()
    if mode in ("product", "version"):
        product_part, version_split = split_product_version(text)
        text = (product_part if mode == "product" else version_split)
    elif sep and part is not None:
        parts = text.split(sep)
        text = parts[part].strip() if 0 <= part < len(parts) else ""
    return text or None


def _load_sheet(path: str, sheet: str | None):
    """엑셀 → 값 grid. 병합 셀은 좌상단 값을 범위 전체에 forward-fill.

    원본 파일/워크북은 절대 수정하지 않는다(저장 호출 없음, grid 는 별도 메모리 리스트).
    병합 정보 접근을 위해 read_only 는 끈다(자산대장 규모면 무방).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        # 세로/가로 병합 모두 좌상단 값으로 채움(원본 불변, grid 에만 반영).
        for rng in list(ws.merged_cells.ranges):
            r0, c0, r1, c1 = rng.min_row - 1, rng.min_col - 1, rng.max_row - 1, rng.max_col - 1
            if r0 < 0 or r0 >= len(grid) or c0 >= len(grid[r0]):
                continue
            val = grid[r0][c0]
            for r in range(r0, min(r1 + 1, len(grid))):
                for cc in range(c0, c1 + 1):
                    if cc < len(grid[r]):
                        grid[r][cc] = val
        title = ws.title
    finally:
        wb.close()
    return title, grid


def detect_header_row(grid: list[list], max_scan: int = 15) -> int:
    """헤더로 가장 그럴듯한 행 인덱스(0-기반) 추정.

    필드 별칭 매칭 수 + 텍스트 밀도가 높고, 숫자/날짜(데이터 냄새)가 적은 행을 고른다.
    제목행(텍스트 1개)·데이터행(숫자 다수)을 자연스럽게 배제.
    """
    all_hints = [h for hints in FIELD_HINTS.values() for h in hints]
    best_idx, best_score = 0, -10 ** 9
    for i, row in enumerate(grid[:max_scan]):
        texts = [str(c).strip().lower() for c in row if c not in (None, "")]
        if not texts:
            continue
        hint_hits = sum(1 for t in texts if any(h in t for h in all_hints))
        numericish = sum(1 for t in texts if t.replace(".", "").replace("-", "").replace("/", "").isdigit())
        score = hint_hits * 3 + len(texts) - numericish * 2
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    """헤더명 유사도 기반 추천 매핑 {field: letter}."""
    suggestion: dict[str, str] = {}
    used: set[str] = set()
    for field, hints in FIELD_HINTS.items():
        for idx, h in enumerate(headers):
            letter = col_letter(idx)
            if letter in used:
                continue
            hl = str(h or "").strip().lower()
            if hl and any(hint in hl for hint in hints):
                suggestion[field] = letter
                used.add(letter)
                break
    return suggestion


def combine_headers(grid: list[list], hidx: int, header_rows: int) -> list[str]:
    """다단(여러 줄) 헤더를 컬럼별 합성 라벨로 결합 (§다단 헤더).

    병합 forward-fill 후 상위 라벨이 컬럼 전체에 채워지므로, hidx 부터 header_rows 행의
    서로 다른 비어있지 않은 라벨을 ' / '로 잇는다. 예: "제품" + "OS" → "제품 / OS".
    """
    n = max(1, header_rows)
    width = max((len(grid[i]) for i in range(hidx, min(hidx + n, len(grid)))), default=0)
    out = []
    for c in range(width):
        parts: list[str] = []
        for r in range(hidx, min(hidx + n, len(grid))):
            v = grid[r][c] if c < len(grid[r]) else None
            s = "" if v is None else str(v).strip()
            if s and s not in parts:
                parts.append(s)
        out.append(" / ".join(parts))
    return out


def preview(path: str, sheet: str | None, header_row: int | None = None,
            header_rows: int = 1, sample_n: int = 5) -> dict:
    """미리보기. header_row(1-기반) 미지정 시 자동 감지. header_rows>1 이면 다단 헤더 결합."""
    title, grid = _load_sheet(path, sheet)
    if not grid:
        return {"sheet": title, "total_rows": 0, "columns": [], "suggested_mapping": {},
                "header_row": 1, "detected_header_row": 1, "header_rows": 1, "preview_rows": []}
    detected = detect_header_row(grid)
    hidx = (header_row - 1) if header_row else detected
    hidx = max(0, min(hidx, len(grid) - 1))
    headers = combine_headers(grid, hidx, header_rows)
    body = grid[hidx + max(1, header_rows):]
    columns = []
    for idx, header in enumerate(headers):
        samples = []
        for r in body[:sample_n]:
            val = r[idx] if idx < len(r) else None
            if val not in (None, ""):
                samples.append(str(val))
        columns.append({"letter": col_letter(idx), "header": header, "samples": samples})
    return {
        "sheet": title,
        "header_row": hidx + 1,                      # 적용된 헤더 시작행(1-기반)
        "header_rows": max(1, header_rows),
        "detected_header_row": detected + 1,         # 자동 감지값
        "preview_rows": [[("" if c is None else str(c)) for c in r] for r in grid[:12]],  # 헤더행 선택용
        "total_rows": len(body),
        "columns": columns,
        "suggested_mapping": suggest_mapping(headers),
    }


def commit(
    db: Session,
    path: str,
    sheet: str | None,
    mapping: dict[str, str],
    import_batch_id: int,
    mode: str = "append",
    on_warning: str = "skip",
    header_row: int | None = None,
    header_rows: int = 1,
    create_departments: bool = True,
) -> dict:
    """매핑 적용·적재. 반환 {committed, warnings, total_rows}. 병합 채움 + (다단)헤더행 적용."""
    title, grid = _load_sheet(path, sheet)
    hidx = ((header_row - 1) if header_row else detect_header_row(grid)) if grid else 0
    hidx = max(0, min(hidx, len(grid) - 1)) if grid else 0
    span = max(1, header_rows)
    body = grid[hidx + span:] if grid else []
    row_offset = hidx + span + 1  # 엑셀 실제 행번호(헤더 다음 행부터)

    depts = {d.name: d for d in db.scalars(select(Department))}
    specs = {field: parse_spec(value) for field, value in mapping.items()}
    specs = {field: s for field, s in specs.items() if s}

    def cell(row, field):
        return resolve_cell(row, specs.get(field))

    if mode == "replace":
        # 전체 교체: 기존 자산 제거(데모 정책). 운영은 부서/배치 단위 교체 권장.
        for a in db.scalars(select(Asset)):
            db.delete(a)
        db.flush()

    warnings: list[dict] = []
    created_departments: list[str] = []
    committed = 0
    for rownum, row in enumerate(body, start=row_offset):  # 엑셀 실제 행번호
        if all(c in (None, "") for c in row):
            continue
        dept_name = cell(row, "department")
        product_raw = cell(row, "product_key")
        version_raw = cell(row, "version_norm")
        # 제품키 정규화용 제품명 분리.
        product_part, _ = split_product_version(product_raw) if product_raw else ("", "")
        # 버전 폴백은 제품 컬럼의 '원본 셀'(분리 전)에서 추출 → 제품을 pv로 이미 나눴어도 버전 보존.
        prod_spec = specs.get("product_key")
        raw_os_cell = resolve_cell(row, (prod_spec[0], "", None, "")) if prod_spec else None
        _, version_split = split_product_version(raw_os_cell) if raw_os_cell else ("", "")
        version_eff = version_raw or version_split  # 명시 버전 우선, 없으면 원본 셀에서 분리한 버전

        issues = []
        if not dept_name:
            issues.append(("MISSING_DEPARTMENT", ""))
        elif dept_name not in depts:
            if create_departments:
                new_dept = Department(name=dept_name)
                db.add(new_dept)
                db.flush()
                depts[dept_name] = new_dept
                created_departments.append(dept_name)
            else:
                issues.append(("UNKNOWN_DEPARTMENT", dept_name))
        if not product_raw:
            issues.append(("MISSING_PRODUCT", ""))
        if not version_eff:
            issues.append(("MISSING_VERSION", ""))

        if issues:
            for code, val in issues:
                warnings.append({"row": rownum, "issue": code, "value": val})
            required_issue = any(
                code in ("MISSING_DEPARTMENT", "UNKNOWN_DEPARTMENT", "MISSING_PRODUCT")
                for code, _ in issues
            )
            if on_warning == "reject" and required_issue:
                continue  # reject 정책: 필수항목(부서·제품) 누락 행만 거부. 버전은 선택이라 경고만.
            if any(code in ("MISSING_DEPARTMENT", "UNKNOWN_DEPARTMENT") for code, _ in issues):
                continue  # 부서 미상이면 적재 불가(매칭/발송 필수)

        asset_no = cell(row, "asset_no") or f"AUTO-{import_batch_id}-{rownum}"
        # 매핑되지 않은 컬럼 → extra (분할 매핑의 원본 컬럼도 '사용됨'으로 제외)
        used_idx = {s[0] for s in specs.values()}
        extra = {
            col_letter(i): str(row[i]).strip()
            for i in range(len(row))
            if i not in used_idx and row[i] not in (None, "")
        }

        existing = db.scalar(select(Asset).where(Asset.asset_no == asset_no))
        if existing is None:
            asset = Asset(asset_no=asset_no)
            db.add(asset)
        else:
            asset = existing  # upsert(최신 배치 우선)
        asset.department_id = depts[dept_name].id
        asset.product_raw = product_raw
        asset.product_key = normalize_product(product_part or product_raw)
        asset.version_raw = version_eff
        asset.version_norm = normalize_version(version_eff)
        asset.ip = cell(row, "ip")
        asset.owner_name = cell(row, "owner_name")
        asset.owner_team = cell(row, "owner_team")
        asset.owner_contact = cell(row, "owner_contact")
        asset.status = enums.AssetStatus.NORMAL
        asset.import_batch_id = import_batch_id
        asset.extra = extra or None
        committed += 1

    db.flush()
    return {"committed": committed, "warnings": warnings, "total_rows": len(body),
            "created_departments": created_departments}
