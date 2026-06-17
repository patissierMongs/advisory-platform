"""조치 결과 보고서 자동 생성 (§★★★★★) — 국정원/국토부 제출용.

· Excel(.xlsx) : openpyxl — 표 형태 제출 산출물.
· HTML        : 한글 완벽 렌더 + 브라우저 인쇄(Ctrl+P)로 PDF 저장(폰트 의존 없는 포터블 PDF 경로).
"""
from __future__ import annotations

import html
import io
from datetime import date, datetime

from sqlalchemy import select
from sqlalchemy.orm import Session

from .. import enums
from ..models import Advisory, Match
from . import remediation


def _active_matches(db: Session, advisory_id: int) -> list[Match]:
    return db.scalars(
        select(Match).where(Match.advisory_id == advisory_id, Match.status == enums.MatchStatus.MATCHED)
    ).all()


def _cve_rows(advisory: Advisory) -> list[dict]:
    out = []
    for ac in advisory.cves:
        c = ac.cve
        if not c:
            continue
        av = c.affected_versions
        ver = ", ".join(av) if isinstance(av, list) else (av.get("lt", "") + " 미만" if isinstance(av, dict) and "lt" in av else "전체")
        out.append({"cve": c.cve_id, "product": c.product_name, "severity": enums.SEVERITY_KO.get(c.severity, c.severity.value),
                    "versions": ver, "desc": c.description, "source": c.source})
    return out


def build_excel(db: Session, advisory: Advisory) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rem = remediation.advisory_remediation(db, advisory)
    matches = _active_matches(db, advisory.id)

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="0F2742")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=14, color="15294A")

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def widths(ws, ws_widths):
        for i, w in enumerate(ws_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── 시트1: 요약 ──
    ws = wb.active
    ws.title = "조치결과 요약"
    ws["A1"] = "보안권고문 조치 결과 보고서"
    ws["A1"].font = title_font
    ws.append([])
    summary = [
        ("문서번호", advisory.doc_no or "-"),
        ("제목", advisory.title or "-"),
        ("발신기관", advisory.source_org or "-"),
        ("접수일", str(advisory.received_at or "-")),
        ("조치기한", str(advisory.due_at or "-")),
        ("SLA 상태", {"NORMAL": "정상", "IMMINENT": "임박", "OVERDUE": "기한초과", "DONE": "완료"}.get(rem["sla_status"], rem["sla_status"])),
        ("대상 부서", f"{rem['dept_total']}개"),
        ("조치 완료율", f"{rem['done_rate']}%  (완료 {rem['done']} / 진행중 {rem['in_progress']} / 불가 {rem['unable']} / 미회신 {rem['none']})"),
        ("대상 자산", f"{rem['asset_total']}대 (조치완료 {rem['asset_done']}대)"),
        ("보고서 생성", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for k, v in summary:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    widths(ws, [16, 90])

    # ── 시트2: 부서별 조치 현황 ──
    ws2 = wb.create_sheet("부서별 조치현황")
    cols2 = ["부서", "대상자산", "채널", "회신상태", "회신담당", "회신일시", "증빙", "비고"]
    ws2.append(cols2)
    style_header(ws2, 1, len(cols2))
    for d in rem["departments"]:
        ws2.append([
            d["department"], d["asset_count"], " · ".join(d["channels"] or []),
            d["ack_status_ko"], d["ack_by"] or "", (d["ack_updated_at"] or "")[:16].replace("T", " "),
            d["evidence"] or "", d["ack_note"] or "",
        ])
    widths(ws2, [20, 9, 16, 12, 12, 18, 22, 40])

    # ── 시트3: 자산 상세 ──
    ws3 = wb.create_sheet("자산 상세")
    cols3 = ["CVE", "자산번호", "부서", "제품/OS", "버전", "IP", "담당자", "심각도"]
    ws3.append(cols3)
    style_header(ws3, 1, len(cols3))
    for m in matches:
        a, c = m.asset, m.advisory_cve.cve
        ws3.append([m.advisory_cve.cve_id_text, a.asset_no, a.department.name if a.department else "",
                    a.product_raw or a.product_key, a.version_raw or a.version_norm, a.ip or "",
                    a.owner_name or "", enums.SEVERITY_KO.get(c.severity, "") if c else ""])
    widths(ws3, [16, 12, 18, 20, 12, 14, 12, 9])

    # ── 시트4: CVE ──
    ws4 = wb.create_sheet("취약점(CVE)")
    cols4 = ["CVE", "제품", "심각도", "영향버전", "설명", "출처"]
    ws4.append(cols4)
    style_header(ws4, 1, len(cols4))
    for r in _cve_rows(advisory):
        ws4.append([r["cve"], r["product"], r["severity"], r["versions"], r["desc"], r["source"]])
    widths(ws4, [16, 24, 9, 18, 30, 14])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_html(db: Session, advisory: Advisory) -> str:
    rem = remediation.advisory_remediation(db, advisory)
    matches = _active_matches(db, advisory.id)
    e = html.escape
    sla_ko = {"NORMAL": "정상", "IMMINENT": "임박", "OVERDUE": "기한초과", "DONE": "완료"}.get(rem["sla_status"], rem["sla_status"])

    dept_rows = "".join(
        f"<tr><td>{e(d['department'] or '')}</td><td class=c>{d['asset_count']}</td>"
        f"<td>{e(' · '.join(d['channels'] or []))}</td>"
        f"<td class=c><span class='b {_ack_cls(d['ack_status'])}'>{e(d['ack_status_ko'])}</span></td>"
        f"<td>{e(d['ack_by'] or '')}</td><td>{e((d['ack_updated_at'] or '')[:16].replace('T',' '))}</td>"
        f"<td>{e(d['evidence'] or '')}</td><td>{e(d['ack_note'] or '')}</td></tr>"
        for d in rem["departments"]
    ) or "<tr><td colspan=8 class=c>발송 내역 없음</td></tr>"

    asset_rows = "".join(
        f"<tr><td class=mono>{e(m.advisory_cve.cve_id_text)}</td><td class=mono>{e(m.asset.asset_no)}</td>"
        f"<td>{e(m.asset.department.name if m.asset.department else '')}</td>"
        f"<td>{e(m.asset.product_raw or m.asset.product_key)}</td><td>{e(m.asset.version_raw or m.asset.version_norm or '')}</td>"
        f"<td class=mono>{e(m.asset.ip or '')}</td><td>{e(m.asset.owner_name or '')}</td></tr>"
        for m in matches
    )
    cve_rows = "".join(
        f"<tr><td class=mono>{e(r['cve'])}</td><td>{e(r['product'] or '')}</td><td class=c>{e(r['severity'])}</td>"
        f"<td>{e(r['versions'])}</td><td>{e(r['desc'] or '')}</td></tr>"
        for r in _cve_rows(advisory)
    )

    return f"""<!DOCTYPE html><html lang=ko><head><meta charset=utf-8>
<title>조치결과보고서 {e(advisory.doc_no or '')}</title>
<style>
@page{{margin:18mm}}
body{{font-family:'Malgun Gothic','맑은 고딕',-apple-system,sans-serif;color:#1a2433;font-size:12px;line-height:1.55;margin:24px}}
h1{{font-size:20px;color:#15294a;border-bottom:3px solid#0f2742;padding-bottom:8px;margin:0 0 4px}}
.sub{{color:#6b7787;font-size:12px;margin-bottom:18px}}
h2{{font-size:14px;color:#0f5d56;margin:22px 0 6px;border-left:4px solid#0f766e;padding-left:8px}}
table{{width:100%;border-collapse:collapse;margin:6px 0;font-size:11px}}
th,td{{border:1px solid#cdd8e6;padding:5px 7px;text-align:left}}
th{{background:#0f2742;color:#fff;font-weight:700}}
td.c{{text-align:center}} .mono{{font-family:Consolas,monospace}}
.kv td:first-child{{background:#f4f6f9;font-weight:700;width:130px}}
.b{{font-size:10px;font-weight:700;padding:1px 8px;border-radius:4px}}
.done{{background:#e6f4f2;color:#0f766e}} .prog{{background:#fff3e2;color:#b45309}}
.none{{background:#fdecec;color:#b42318}} .unable{{background:#eef1f5;color:#51607a}}
.rate{{font-size:26px;font-weight:800;color:#0f766e}}
@media print{{body{{margin:0}} .b,th{{print-color-adjust:exact;-webkit-print-color-adjust:exact}}}}
</style></head><body>
<h1>보안권고문 조치 결과 보고서</h1>
<div class=sub>생성 {datetime.now().strftime('%Y-%m-%d %H:%M')} · 본 보고서는 [보안권고문 처리 시스템]에서 자동 생성되었습니다.</div>
<table class=kv>
<tr><td>문서번호</td><td>{e(advisory.doc_no or '-')}</td></tr>
<tr><td>제목</td><td>{e(advisory.title or '-')}</td></tr>
<tr><td>발신기관 / 접수일</td><td>{e(advisory.source_org or '-')} / {e(str(advisory.received_at or '-'))}</td></tr>
<tr><td>조치기한</td><td>{e(str(advisory.due_at or '-'))} (D{rem['d_day']:+d}, {e(sla_ko)})</td></tr>
</table>
<h2>조치 요약</h2>
<table><tr>
<td style="width:25%;text-align:center"><div class=rate>{rem['done_rate']}%</div>조치 완료율</td>
<td>대상 부서 <b>{rem['dept_total']}</b>개 · 완료 <b>{rem['done']}</b> / 진행중 {rem['in_progress']} / 불가 {rem['unable']} / 미회신 {rem['none']}<br>
대상 자산 <b>{rem['asset_total']}</b>대 (조치완료 {rem['asset_done']}대)</td>
</tr></table>
<h2>부서별 조치 현황</h2>
<table><tr><th>부서</th><th>대상자산</th><th>채널</th><th>회신상태</th><th>담당</th><th>회신일시</th><th>증빙</th><th>비고</th></tr>{dept_rows}</table>
<h2>취약점(CVE)</h2>
<table><tr><th>CVE</th><th>제품</th><th>심각도</th><th>영향버전</th><th>설명</th></tr>{cve_rows}</table>
<h2>대상 자산 상세</h2>
<table><tr><th>CVE</th><th>자산번호</th><th>부서</th><th>제품/OS</th><th>버전</th><th>IP</th><th>담당자</th></tr>{asset_rows}</table>
</body></html>"""


def _ack_cls(ack: str) -> str:
    return {"DONE": "done", "IN_PROGRESS": "prog", "NONE": "none", "UNABLE": "unable"}.get(ack, "none")
